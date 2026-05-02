"""Leitura de documentos para alimentar o pipeline de estruturação.

Cobre os formatos hoje suportados pelo backend e pelo SDK:
``.txt``, ``.md``, ``.docx``, ``.csv``, ``.xlsx``. A função pública é
:func:`read_document`, que recebe ``bytes`` + ``filename`` (mesmo
contrato que o serviço usa para arquivos baixados do Blob Storage).

Cada arquivo vira ``list[tuple[doc_id, texto]]``. Para ``.txt/.md/.docx``
isso é uma lista de um item; para ``.csv/.xlsx`` é uma lista por linha.

Imports de ``openpyxl`` e do parser DOCX são lazy — o caller que pede
``.xlsx`` é quem precisa ter a dep instalada (``labdados-core[estruturacao]``).
"""

from __future__ import annotations

import csv
import io
import os
import zipfile
from typing import Any

SUPPORTED_EXTENSIONS = (".txt", ".md", ".docx", ".csv", ".xlsx")


def read_document(
    content: bytes,
    filename: str,
    *,
    csv_text_column: str = "",
) -> list[tuple[str, str]]:
    """Devolve ``[(doc_id, texto), ...]`` extraído do arquivo.

    Para ``.csv/.xlsx``, ``csv_text_column`` indica a coluna que contém
    o texto a estruturar; vazio = concatena todas as colunas
    (``"chave: valor"`` por linha).
    """
    ext = os.path.splitext(filename)[1].lower()
    stem = os.path.splitext(filename)[0]

    if ext == ".docx":
        return [(stem, _read_docx(content))]
    if ext == ".csv":
        return _read_csv_rows(content, stem, csv_text_column)
    if ext == ".xlsx":
        return _read_xlsx_rows(content, stem, csv_text_column)
    if ext in (".txt", ".md", ""):
        return [(stem, content.decode("utf-8", errors="replace"))]
    raise ValueError(f"Extensão não suportada: {ext!r} (suportadas: {SUPPORTED_EXTENSIONS})")


def _read_docx(content: bytes) -> str:
    """Extrai texto de .docx (ZIP de XML) sem dependência externa."""
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        if "word/document.xml" not in zf.namelist():
            return ""
        xml_content = zf.read("word/document.xml")
        tree = ET.fromstring(xml_content)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = []
        for p in tree.iter(f"{{{ns['w']}}}p"):
            texts = [t.text or "" for t in p.findall(".//w:t", ns)]
            paragraphs.append("".join(texts))
        return "\n".join(paragraphs)


def _row_to_text(row: dict[str, Any], text_column: str) -> str:
    if text_column:
        target = text_column.strip().lower()
        for k, v in row.items():
            if str(k).strip().lower() == target:
                return "" if v is None else str(v)
        return ""
    parts = [f"{k}: {v}" for k, v in row.items() if v not in (None, "")]
    return "\n".join(parts)


def _read_csv_rows(content: bytes, stem: str, text_column: str) -> list[tuple[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    docs: list[tuple[str, str]] = []
    for idx, row in enumerate(reader, start=1):
        doc_text = _row_to_text(row, text_column)
        docs.append((f"{stem}__row{idx:04d}", doc_text))
    return docs


def _read_xlsx_rows(content: bytes, stem: str, text_column: str) -> list[tuple[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Leitura de .xlsx requer openpyxl. Instale `labdados-core[estruturacao]`."
        ) from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows)]
    except StopIteration:
        return []
    docs: list[tuple[str, str]] = []
    for idx, raw_row in enumerate(rows, start=1):
        row_dict = {header[i]: raw_row[i] if i < len(raw_row) else None for i in range(len(header))}
        doc_text = _row_to_text(row_dict, text_column)
        docs.append((f"{stem}__row{idx:04d}", doc_text))
    return docs
